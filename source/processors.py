# processors.py
import pandas as pd
import pickle
from text_classification import classify_text
from pdf_table_grid import TableExtractorApp
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
import re
import os
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import column_index_from_string
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection
from datetime import datetime
from openpyxl.worksheet.datavalidation import DataValidation
from os.path import join
import numpy as np
from copy import copy
from excel_format import prepare_tables_for_writing, write_df_to_named_table_by_header

PATH_MODEL = join('..', 'prediccion')

columns_to_extract = {'Cuenta BCI 18' : ['Fecha', 'Sucursal', 'Descripción', 'N° Documento', 'Cheques y otros cargos' , 'Depósitos y Abono', 'Saldo diario'],
                      'Cuenta BCI 85' : ['Fecha', 'Sucursal', 'Descripción', 'N° Documento', 'Cheques y otros cargos' , 'Depósitos y Abono', 'Saldo diario'],
                      'Siteminder' : ['Referencia de la reserva', 'Nombres de los huéspedes', 'Llegada', 'Salida', 'Canal', 'Affiliated Channel', 'Referral', 'Habitación', 'Fecha de reserva', 'Estado de la reserva',	'Ocupación', 'Precio total'],
                      'Tarjeta de crédito internacional' : ['NUMERO REFERENCIA INTERNACIONAL', 'FECHA OPERACIÓN', 'DESCRIPCION OPERACION O COBRO', 'CIUDAD', 'PAIS', 'MONTO MONEDA ORIGEN', 'MONTO US$'],
                      'Tarjeta de crédito nacional 24' : ['LUGAR DE OPERACIÓN', 'FECHA OPERACIÓN', 'CODIGO REFERENCIA', 'DESCRIPCION OPERACION O COBRO', 'MONTO OPERACIÓN O COBRO', 'MONTO TOTAL A PAGAR'],
                      'Tarjeta de crédito nacional 69' : ['LUGAR DE OPERACIÓN', 'FECHA OPERACIÓN', 'CODIGO REFERENCIA', 'DESCRIPCION OPERACION O COBRO', 'MONTO OPERACIÓN O COBRO', 'MONTO TOTAL A PAGAR'],
                      'Cuenta BCI Comercio Exterior' : ['FECHA', 'SUCURSAL', 'DESCRIPCION', 'N° DE DOCUMENTO', 'CHEQUES Y OTROS CARGOS', 'DEPOSITOS Y ABONOS', 'SALDO DIARIO'],
                      'Banco Security' : ['fecha ', 'descripción ', 'número de documentos ', 'cargos ', 'abonos ', 'saldos '],
                      'Transbank' : ['Fecha de venta - nan', 'Cód. comercio - nan',
                                      'Nombre local - nan', 'Tipo de movimiento - nan',
                                      'Tipo de tarjeta - nan', 'Identificador - N° de tarjeta',
                                      'nan - Cód. de autorización', 'nan - Orden de pedido',
                                      'nan - Número único', 'nan - ID de servicio', 'Tipo de cuota - nan',
                                      'Monto afecto - nan', 'Monto exento - nan', 'N° de cuotas - nan',
                                      'Monto cuota - nan', 'Fecha de abono - nan', 'N° de boleta - nan',
                                      'Monto vuelto - nan'],
                      'Transferencias' : ['Monto $', 'Nombre Destino/Origen', 'Mensaje Destino']}

column_dictionary = {'Banco Security' : {'fecha ' : 'FECHA', 
                                   'descripción ' : 'DESCRIPCION', 
                                   'número de documentos ' : 'Nº DOCUMENTO', 
                                   'cargos ' : 'CARGO', 
                                   'abonos ' : 'ABONO', 
                                   'saldos ' : 'SALDO'},
                     'Siteminder' : {'Referencia de la reserva' : 'ID RESERVA', 
                                     'Nombres de los huéspedes' : 'HUESPEDES', 
                                     'Llegada' : 'LLEGADA', 
                                     'Salida' : 'SALIDA',
                                     'Canal' : 'CANAL', 
                                     'Affiliated Channel' :'CANAL AFILIADO', 
                                     'Estado de la reserva' : 'ESTADO', 
                                     #'Ocupación' : 'Habitaciones', 
                                     'Habitación' : 'HABITACIONES', 
                                     'Precio total' : 'TOTAL',
                                     'Fecha de reserva' : 'FECHA RESERVA'},
                     'Transbank' : {'Fecha de venta - nan' : 'FECHA', 
                                    'Cód. comercio - nan' : 'CODIGO COMERCIO',
                                    'Nombre local - nan' : 'LOCAL',
                                    'Tipo de movimiento - nan' : 'TIPO MOVIMIENTO',
                                    'Tipo de tarjeta - nan': 'TARJETA',
                                    'Identificador - N° de tarjeta' : 'NUMERO TARJETA',
                                    'nan - Cód. de autorización' : 'CODIGO AUTORIZACION',
                                    'nan - Orden de pedido' : 'ORDEN',
                                    'nan - Número único' : 'NUMERO OPERACION',
                                    'nan - ID de servicio' : 'ID SERVICIO',
                                    'Tipo de cuota - nan' : 'TIPO CUOTA',
                                    'Monto afecto - nan' : 'TOTAL',
                                    'Monto exento - nan' : 'MONTO',
                                    'N° de cuotas - nan' : 'NUMERO CUOTAS',
                                    'Monto cuota - nan' : 'MONTO CUOTA',
                                    'Fecha de abono - nan' : 'FECHA ABONO',
                                    'N° de boleta - nan' : 'NUMERO BOLETA',
                                    'Monto vuelto - nan' : 'VUELTO'},
                     'Transferencias' : {'Monto $' : 'CARGO TRANSF', 
                                         'Nombre Destino/Origen' : 'NOMBRE DESTINO', 
                                         'Mensaje Destino' : 'MENSAJE'},
                     'Cuenta BCI Comercio Exterior' : {'SUCURSAL' : 'OFICINA', 
                                                       'CHEQUES Y OTROS CARGOS' : 'CARGO', 
                                                       'DEPOSITOS Y ABONOS' : 'ABONO', 
                                                       'SALDO DIARIO' : 'SALDO'},
                     'Cuenta BCI 85' : {'Fecha' : 'FECHA', 
                                        'Sucursal' : 'OFICINA', 
                                        'Descripción' : 'MOVIMIENTO', 
                                        'N° Documento': 'N° DOCUMENTO', 
                                        'Cheques y otros cargos' : 'CARGO',
                                        'Depósitos y Abono' : 'ABONO',
                                        'Saldo diario' : 'SALDO'},
                     'Cuenta BCI 18' : {'Fecha' : 'FECHA', 
                                        'Sucursal' : 'OFICINA', 
                                        'Descripción' : 'DESCRIPCION', 
                                        'N° Documento': 'Nº DOCUMENTO', 
                                        'Cheques y otros cargos' : 'CARGO',
                                        'Depósitos y Abono' : 'ABONO',
                                        'Saldo diario' : 'SALDO'},
                     'Tarjeta de crédito nacional 24' : {'FECHA OPERACIÓN' : 'FECHA', 
                                                         'CODIGO REFERENCIA' : 'CÓDIGO REFERENCIA', 
                                                         'LUGAR DE OPERACIÓN' : 'LUGAR OPERACIÓN', 
                                                         'DESCRIPCION OPERACION O COBRO': 'DESCRIPCION', 
                                                         'MONTO OPERACIÓN O COBRO' : 'CARGO'},
                     'Tarjeta de crédito nacional 69' : {'FECHA OPERACIÓN' : 'FECHA', 
                                                         'CODIGO REFERENCIA' : 'CÓDIGO REFERENCIA', 
                                                         'LUGAR DE OPERACIÓN' : 'LUGAR OPERACIÓN', 
                                                         'DESCRIPCION OPERACION O COBRO': 'DESCRIPCION', 
                                                         'MONTO OPERACIÓN O COBRO' : 'CARGO'},
                     'Tarjeta de crédito internacional' : {'FECHA OPERACIÓN' : 'FECHA', 
                                                           'NUMERO REFERENCIA INTERNACIONAL' : 'CÓDIGO REFERENCIA', 
                                                           'DESCRIPCION OPERACION O COBRO' : 'DESCRIPCION'}}

def convert_date(row):
    #return datetime.strptime(row, "%m/%d/%Y")
    return datetime.strptime(row, "%Y-%m-%d")

def parse_number(value, label):
    if isinstance(value, (int, float)):
        return abs(value)

    if not isinstance(value, str):
        return None

    value = value.strip()
    if not value:
        return None

    # Eliminar cualquier signo negativo y espacios como "- 5.000.000"
    value = value.lstrip('-').replace(' ', '')

    # General cleaning: eliminar símbolos monetarios y letras
    value = re.sub(r'[^\d.,]', '', value)

    # ----- FORMATO POR DOCUMENTO -----
    if label in ['Cuenta BCI 85', 'Cuenta BCI 18', 'Tarjeta de crédito nacional 24', 'Tarjeta de crédito nacional 69']:
        # Ejemplo: 5.050.786 → sin decimales, con separador de miles '.'
        value = value.replace('.', '')
    
    elif label in ['Cuenta BCI Comercio Exterior', 'Tarjeta de crédito internacional', 'Transbank']:
        # Ejemplo: 116.957,02 → miles con '.' y decimales con ','
        value = value.replace('.', '').replace(',', '.')

    elif label == 'Transferencias':
        # Ejemplo: - 5.000.000 → miles con '.', sin decimales, con espacio y signo
        value = value.replace('.', '')

    elif label == 'Siteminder':
        # Ejemplo: 1433.61 → sin separador de miles, decimales con punto
        value = value.replace(',', '')  # por si acaso

    elif label == 'Security':
        # Ejemplo: 2330,51 → sin miles, decimales con coma
        value = value.replace(',', '.')

    else:
        # Fallback: intenta autodetectar
        if re.match(r'^\d{1,3}(\.\d{3})*,\d{1,2}$', value):
            value = value.replace('.', '').replace(',', '.')
        elif ',' in value and '.' not in value:
            value = value.replace(',', '.')
        else:
            value = value.replace(',', '')

    try:
        float_val = float(value)
        return int(float_val) if float_val.is_integer() else float_val
    except ValueError:
        return None

def apply_classification(df, model_path, vectorizer_path, text_columns):
    """
    Apply a classification model to a dataframe using specified text columns.

    Parameters:
        df (pd.DataFrame): The input dataframe.
        model_path (str): Path to the .pkl file for the trained model.
        vectorizer_path (str): Path to the .pkl file for the vectorizer.
        text_columns (list): List of column names whose values are used as input to the model.

    Returns:
        pd.DataFrame: The original dataframe with two new columns:
                      ['Predicción Clave', 'Puntaje Confianza']
    """
    # Load model and vectorizer
    with open(model_path, "rb") as model_file, open(vectorizer_path, "rb") as vectorizer_file:
        model = pickle.load(model_file)
        vectorizer = pickle.load(vectorizer_file)

    # Ensure only available columns are used
    available_cols = [col for col in text_columns if col in df.columns]

    def classify_row(row):
        inputs = [str(row[col]) for col in available_cols if pd.notnull(row[col])]
        return pd.Series(classify_text(model, vectorizer, *inputs))

    df[['CLAVE', 'PREDICCIÓN']] = df.apply(classify_row, axis=1)
    return df

def file_to_df(path, label):
    if label == 'Banco Security':
        header = 14
        df = pd.read_excel(path, header=header - 1, thousands='.')
        end_index = None
        # Itera hasta encontrar un espacio y para
        for i, row in df.iterrows():
            if pd.isna(row[0]):
                end_index = i
                break 
        df = df.iloc[: end_index]

    elif label == 'Transbank':
        df = process_transbank_directory(path)
    
    elif label == 'Siteminder':
        df = pd.read_csv(path)

    elif label in ('Cuenta BCI Comercio Exterior', 'Tarjeta de crédito nacional 24', 'Tarjeta de crédito nacional 69', 'Tarjeta de crédito internacional'):
        # Aquí se abre tu GUI personalizada que devuelve un DataFrame
        df = launch_pdf_gui(path, label)  # Esta GUI procesa el PDF y devuelve df

    elif label == 'Cuenta BCI 18':
        header = 18 # Fila donde están los nombres de la columna
        df = pd.read_excel(path, header=header - 1, thousands='.')

    elif label == 'Cuenta BCI 85':
        header = 23 # Fila donde están los nombres de la columna (Numero de fila en excel - 1)
        df = pd.read_excel(path, header=header - 1, thousands='.')

    elif label == 'Transferencias':
        engine = 'xlrd' # El motor openpyxl no usa xls
        header = 9 # Fila donde están los nombres de la columna (Numero de fila en excel - 1)
        df = pd.read_excel(path, header=header - 1, thousands='.')

    df = df[columns_to_extract[label]]
    df.rename(columns=column_dictionary[label], inplace=True)
    df = format_column(df, label)

    return df

def convert_numeric(value):
    value = value.replace('.', '')
    numeric = float(value.split()[1])
    return numeric

def format_column(df, label):
    if label in ['Cuenta BCI 85', 'Cuenta BCI 18']:
        numeric_cols = ['CARGO', 'ABONO', 'SALDO']
        for col in numeric_cols:
            df[col] = df[col].apply(lambda x: parse_number(x, label))

    elif label == 'Cuenta BCI Comercio Exterior':
        numeric_cols = ['CARGO', 'ABONO', 'SALDO']
        for col in numeric_cols:
            df[col] = df[col].apply(lambda x: parse_number(x, label))

    elif label == 'Transferencias':
        numeric_cols = ['CARGO TRANSF']
        for col in numeric_cols:
            df[col] = df[col].apply(lambda x: parse_number(x, label))

    elif label in ['Tarjeta de crédito nacional 24', 'Tarjeta de crédito nacional 69']:
        numeric_cols = ['CARGO']
        for col in numeric_cols:
            #df[col] = df[col].apply(lambda x: x.split()[-1] if isinstance(x, str) else x)
            df[col] = df[col].apply(lambda x: parse_number(x, label))

    elif label == 'Tarjeta de crédito internacional':
        numeric_cols = ['MONTO US$']
        for col in numeric_cols:
            df[col] = df[col].apply(lambda x: parse_number(x, label))

    elif label == 'Siteminder':
        # Formato numérico
        numeric_cols = ['TOTAL']
        for col in numeric_cols:
            #df[col] = df[col].apply(lambda x: x.split()[-1] if isinstance(x, str) else x)
            df[col] = df[col].apply(lambda x: parse_number(x, label))
        # Limpiar y convertir columnas
        df['HABITACIONES'] = df['HABITACIONES'].apply(lambda x: x[0])
        df['LLEGADA'] = df['LLEGADA'].apply(convert_date).dt.date
        df['SALIDA'] = df['SALIDA'].apply(convert_date).dt.date

    elif label == 'Banco Security':
        numeric_cols = ['CARGO', 'ABONO', 'SALDO']
        for col in numeric_cols:
            df[col] = df[col].apply(lambda x: parse_number(x, label))

    elif label == 'Transbank':

        # Nuevas columnas
        df['PESOS'] = np.nan
        df['TIPO VENTA'] = np.nan
        df['PROPINA'] = np.nan
        df['CODIGO EMPLEADO'] = np.nan

        # Asignar moneda
        df['MONEDA'] = df['TIPO MOVIMIENTO'].replace({
            'Venta USD$': 'DOLAR',
            'Venta $': 'PESO'
        })

        # Convertir número de boleta
        df['NUMERO'] = pd.to_numeric(df['NUMERO BOLETA'], errors='coerce')

        # Convertir fecha y ordenar
        df[['FECHA', 'HORA']] = df['FECHA'].str.split(' ', n=1, expand=True)
        df['FECHA'] = pd.to_datetime(df['FECHA'], format='%d/%m/%Y', errors='coerce')
        df = df.sort_values(by=['FECHA', 'NUMERO'], ascending=True)

        numeric_cols = ['TOTAL', 'MONTO', 'MONTO CUOTA', 'VUELTO']
        for col in numeric_cols:
            df[col] = df[col].apply(lambda x: parse_number(x, label))

    return df

def process_transbank_directory(directory):
    dataframes = []
    for filename in os.listdir(directory):
        if not filename.lower().endswith(('.xlsx', '.xls')):
            continue

        file_path = os.path.join(directory, filename)

        try:
            df_temp = pd.read_excel(file_path, header=None)
            columnas = df_temp.iloc[17].astype(str).str.strip() + " - " + df_temp.iloc[18].astype(str).str.strip()

            df = pd.read_excel(file_path, skiprows=19, header=None)
            df.columns = columnas

            dataframes.append(df)

        except Exception as e:
            print(f"Error processing {filename}: {e}")
            continue

    if dataframes:
        df_stack = pd.concat(dataframes, ignore_index=True)
        return df_stack
    else:
        return pd.DataFrame()

def df_to_wb(dfs, wb, sheet_name, model_dirpath):
    ws = wb[sheet_name]
    table_data = []  # Will contain tuples (table_name, df)
    if sheet_name == 'BCI':
        # ------------------ Cuenta BCI 85 + Transferencias ------------------
        if 'Transferencias' in dfs and 'Cuenta BCI 85' in dfs:
            df_transfer = dfs['Transferencias'].copy()
            df_85 = dfs['Cuenta BCI 85'].copy()
            desc_list = ['TRASPASO FONDOS OTRO BANCO EN LINEA', 'CARGO POR TRANSF DE FONDOS AUTOSERVICIO']
            mask_desc = df_85['MOVIMIENTO'].isin(desc_list) | df_85['MOVIMIENTO'].str.contains('Transfer', case=False, na=False)
            filtered_cartola = df_85[mask_desc].copy()
            filtered_cartola['merge_id'] = filtered_cartola.groupby(['CARGO']).cumcount()
            df_transfer['merge_id'] = df_transfer.groupby(['CARGO TRANSF']).cumcount()

            merged_cartola = filtered_cartola.merge(
                df_transfer[['CARGO TRANSF', 'NOMBRE DESTINO', 'MENSAJE', 'merge_id']],
                how='left',
                left_on=['CARGO', 'merge_id'],
                right_on=['CARGO TRANSF', 'merge_id']
            )

            non_filtered_cartola = df_85[~mask_desc].copy()
            df_final = pd.concat([merged_cartola, non_filtered_cartola]).sort_index()
            model_path, vectorizer_path = join(model_dirpath, "modelo_C85.pkl"), join(model_dirpath, "vectorizer_C85.pkl")
            df_final = apply_classification(df_final, model_path, vectorizer_path, ['MOVIMIENTO', 'NOMBRE DESTINO', 'MENSAJE'])
            df_final = df_final.drop(columns=['CARGO TRANSF', 'merge_id'], errors='ignore')
            table_data.append(('Cuenta_85', df_final))

        # ------------------ Cuenta BCI Comercio Exterior ------------------
        if 'Cuenta BCI Comercio Exterior' in dfs:
            df_ext = dfs['Cuenta BCI Comercio Exterior'].copy()
            model_path, vectorizer_path = join(model_dirpath, "modelo_CUSD.pkl"), join(model_dirpath, "vectorizer_CUSD.pkl")
            df_ext = apply_classification(df_ext, model_path, vectorizer_path, ['DESCRIPCION'])
            table_data.append(('BCI_Comercio_Exterior', df_ext))

        # ------------------ Tarjeta de crédito nacional ------------------
        if 'Tarjeta de crédito nacional 24' in dfs or 'Tarjeta de crédito nacional 69' in dfs:
            df_tcn_list = [dfs[k] for k in ['Tarjeta de crédito nacional 24', 'Tarjeta de crédito nacional 69'] if k in dfs]
            df_credito_nacional = pd.concat(df_tcn_list, ignore_index=True)
            model_path, vectorizer_path = join(model_dirpath, "modelo_TCN.pkl"), join(model_dirpath, "vectorizer_TCN.pkl")
            df_credito_nacional = apply_classification(df_credito_nacional, model_path, vectorizer_path, ['DESCRIPCION'])
            table_data.append(('TC_Nacional', df_credito_nacional))  # Use the actual table name here

        # ------------------ Tarjeta de crédito internacional ------------------
        if 'Tarjeta de crédito internacional' in dfs:
            df_credito_internacional = dfs['Tarjeta de crédito internacional'].copy()
            model_path, vectorizer_path = join(model_dirpath, "modelo_TCI.pkl"), join(model_dirpath, "vectorizer_TCI.pkl")
            df_credito_internacional = apply_classification(df_credito_internacional, model_path, vectorizer_path, ['DESCRIPCION'])
            df_credito_internacional['DESCRIPCION'] = df_credito_internacional['CIUDAD'].fillna('') + ' ' + df_credito_internacional['DESCRIPCION'].fillna('')
            start_row = 1  # Dummy value for now
            df_credito_internacional['CARGO'] = [
                f"=D{row_num}*EERR!$D$2" for row_num in range(start_row, start_row + len(df_credito_internacional))
            ]
            table_data.append(('TC_Internacional', df_credito_internacional))  # Use actual table name
    elif sheet_name == 'Security':
        if 'Banco Security' in dfs:
            df_security = dfs['Banco Security']
            model_path, vectorizer_path = join(model_dirpath, "modelo_security.pkl"), join(model_dirpath, "vectorizer_security.pkl")
            text_columns = ['DESCRIPCION']
            df_security = apply_classification(df_security, model_path, vectorizer_path, text_columns)
            table_data.append(('Security', df_security))
    elif sheet_name == 'Siteminder':
        if 'Siteminder' in dfs:
            df_siteminder = dfs['Siteminder']
            table_data.append(('Siteminder', df_siteminder))
    elif sheet_name == 'BCI FondRendir':
        if 'Cuenta BCI 18' in dfs:
            df_cta18 = dfs['Cuenta BCI 18']
            table_data.append(('Cuenta_18', df_cta18))
    elif sheet_name == 'Transbank':
        if 'Transbank' in dfs:
            df_trans = dfs['Transbank']
            print(len(df_trans))
            table_data.append(('Transbank', df_trans))

    # ------------------ Apply all changes to sheet ------------------
    if table_data:
        prepare_tables_for_writing(ws, wb, table_data)
        for table_name, df in table_data:
            write_df_to_named_table_by_header(ws, sheet_name, table_name, df)

def launch_pdf_gui(file_path, file_type=None):
    import tkinter as tk
    import pandas as pd
    root = tk.Toplevel()
    extractor = TableExtractorApp(root, file_path=file_path)
    root.grab_set()
    root.wait_window()
    return getattr(extractor, 'result_df', pd.DataFrame())

