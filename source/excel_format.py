from openpyxl.styles import Alignment, Font, PatternFill, Border, Protection
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell.cell import MergedCell
from copy import copy


def reapply_data_validations(ws, table, start_row, num_new_rows):
    min_col, min_row, max_col, _ = range_boundaries(table.ref)
    for dv in list(ws.data_validations.dataValidation):
        for col in range(min_col, max_col + 1):
            ref_cell = f"{get_column_letter(col)}{min_row + 1}"
            if any(ref_cell in str(rng) for rng in dv.ranges):
                new_range = f"{get_column_letter(col)}{start_row + 1}:{get_column_letter(col)}{start_row + num_new_rows}"
                new_dv = copy(dv)
                new_dv.ranges = []
                new_dv.add(new_range)
                ws.add_data_validation(new_dv)

def reapply_data_validations_post_move(ws, original_ref, new_ref, num_new_rows):
    from openpyxl.utils import range_boundaries, get_column_letter

    o_min_col, o_min_row, o_max_col, _ = range_boundaries(original_ref)
    n_min_col, n_min_row, n_max_col, _ = range_boundaries(new_ref)

    for dv in list(ws.data_validations.dataValidation):
        for col in range(o_min_col, o_max_col + 1):
            col_letter = get_column_letter(col)
            for rng in dv.ranges:
                if isinstance(rng, str) and f"{col_letter}{o_min_row + 1}" in rng:
                    new_range = f"{col_letter}{n_min_row + 1}:{col_letter}{n_min_row + num_new_rows}"
                    new_dv = copy(dv)
                    new_dv.ranges = []
                    new_dv.add(new_range)
                    ws.add_data_validation(new_dv)

def write_df_to_named_table_by_header(ws, sheet_name, table_name, df):
    table = ws.tables.get(table_name)
    if table is None:
        raise ValueError(f"Table '{table_name}' not found in sheet '{sheet_name}'.")

    start_cell, _ = table.ref.split(':')
    start_col_letter = ''.join(filter(str.isalpha, start_cell))
    start_col = column_index_from_string(start_col_letter)
    start_row = int(''.join(filter(str.isdigit, start_cell)))

    headers = [ws.cell(row=start_row, column=col).value for col in range(start_col, start_col + len(table.tableColumns))]
    header_to_col = {header: col for header, col in zip(headers, range(start_col, start_col + len(headers)))}

    for row_offset, (_, df_row) in enumerate(df.iterrows(), start=1):
        for header, col_idx in header_to_col.items():
            if header in df.columns:
                cell = ws.cell(row=start_row + row_offset, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = df_row[header]

    new_end_row = start_row + len(df)
    table.ref = f"{start_cell}:{ws.cell(row=new_end_row, column=start_col + len(headers) - 1).coordinate}"
    reapply_data_validations(ws, table, start_row, len(df))

    first_data_row = start_row + 1
    for col_idx in range(start_col, start_col + len(headers)):
        formula = ws.cell(row=first_data_row, column=col_idx).value
        if isinstance(formula, str) and formula.startswith('='):
            for offset in range(1, len(df)):
                ws.cell(row=first_data_row + offset, column=col_idx, value=formula)

    return new_end_row

def move_table_and_label_down(ws, table, offset, label_rows_above=1):
    from openpyxl.worksheet.cell_range import CellRange

    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    label_start_row = max(min_row - label_rows_above, 1)
    full_range = f"{get_column_letter(min_col)}{label_start_row}:{get_column_letter(max_col)}{max_row}"

    # Find merged cells that intersect with the label rows
    merged_to_move = []
    for cr in list(ws.merged_cells.ranges):
        if cr.min_row >= label_start_row and cr.max_row < min_row:
            merged_to_move.append(cr)
            ws.unmerge_cells(str(cr))  # 🔁 Unmerge BEFORE the move

    # Store label cell positions (for clearing styles)
    label_cells = [(row, col) for row in range(label_start_row, min_row) for col in range(min_col, max_col + 1)]

    # Move the entire block (labels + table)
    ws.move_range(full_range, rows=offset, cols=0)

    # Recreate merged cells in the new position
    for cr in merged_to_move:
        new_range = CellRange(
            min_col=cr.min_col,
            min_row=cr.min_row + offset,
            max_col=cr.max_col,
            max_row=cr.max_row + offset,
        )
        ws.merge_cells(str(new_range))

    # Clear formatting from original label cells, but skip merged cells
    for row, col in label_cells:
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            continue
        cell.font = Font()
        cell.fill = PatternFill()
        cell.border = Border()
        cell.alignment = Alignment()
        cell.number_format = 'General'
        cell.protection = Protection()

    # Update the table reference to the new location
    new_ref = f"{get_column_letter(min_col)}{min_row + offset}:{get_column_letter(max_col)}{max_row + offset}"
    table.ref = new_ref

def prepare_tables_for_writing(ws, wb, write_instructions, label_rows_above=1):
    table_map = {t.name: t for t in ws.tables.values()}
    table_lengths = {name: len(df) for name, df in write_instructions if name in table_map}

    def table_start_row(table):
        return range_boundaries(table.ref)[1]

    sorted_tables = sorted(ws.tables.values(), key=table_start_row)
    new_ends = {name: table_start_row(table) + table_lengths.get(name, 0) for name, table in table_map.items()}

    for i in range(len(sorted_tables) - 1):
        upper = sorted_tables[i]
        lower = sorted_tables[i + 1]
        spacing = 3
        upper_end = new_ends[upper.name]
        lower_start = table_start_row(lower)
        needed_offset = (upper_end + spacing) - lower_start

        if needed_offset > 0:
            original_ref = lower.ref  # Save original position BEFORE move
            move_table_and_label_down(ws, lower, needed_offset, label_rows_above)
            new_ends[lower.name] += needed_offset
            new_ref = lower.ref  # New position AFTER move

            # Reapply validations to the moved table's new position
            # Use length from write_instructions if available, else fallback to current length
            num_new_rows = table_lengths.get(lower.name, range_boundaries(new_ref)[3] - range_boundaries(new_ref)[1])
            reapply_data_validations_post_move(ws, original_ref, new_ref, num_new_rows)

    last_table = sorted_tables[-1]
    last_end = new_ends[last_table.name]
    min_col, _, max_col, _ = range_boundaries(last_table.ref)
    for row in range(last_end + 1, last_end + 3):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.strip():
                cell.value = None
                cell.font = Font()
                cell.fill = PatternFill()
                cell.border = Border()
                cell.alignment = Alignment()
                cell.number_format = 'General'
                cell.protection = Protection()
